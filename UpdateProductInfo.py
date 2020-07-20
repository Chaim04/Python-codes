from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb_rawdata = load_workbook('D:\Python\Python\Ivy part time\\Raw data.xlsx')
wb_updatedfile = load_workbook('D:\Python\Python\Ivy part time\\Updated file.xlsx')
ws_Temp_UK = wb_updatedfile["Template_UK"]
ws_Temp_DE = wb_updatedfile["Template_DE"]
ws_Temp_FR = wb_updatedfile["Template_FR"]
ws_Temp_IT = wb_updatedfile["Template_IT"]
ws_Temp_ES = wb_updatedfile["Template_ES"]
ws_Rawdata = wb_rawdata["Rawdata"]

"""list out the product & child_qty"""

Child_qty = []
for cell in ws_Rawdata["2"]:
    col = cell.column
    Child_qty.append(ws_Rawdata.cell(row = 2,column = col).value)

del Child_qty[0]

Product_Qty = len(Child_qty)

"""define the column of these info """

cur_row_UK = 4
cur_row_DE = 4
cur_row_FR = 4
cur_row_IT = 4
cur_row_ES = 4
col_item_sku = 2
col_item_name = 6
col_product_description_UK = 37
col_generic_keywords1_UK = 43
col_bullet_point1_UK = 49
col_bullet_point2_UK = 50
col_bullet_point3_UK = 51
col_bullet_point4_UK = 52
col_bullet_point5_UK = 53
col_standard_price = 26
col_recommended_browse_nodes = 7
col_outer_material_type = 8
col_brand_name = 3
col_external_product_id_type = 5
col_quantity = 27
col_parent_child_UK = 32
col_parent_sku_UK = 33
col_relationship_type_UK = 34
col_material_composition_UK_DE_FR= 13
col_variation_theme_UK = 35
col_update_delete_UK = 36
col_product_description_others = 42
col_generic_keywords1_others = 48
col_bullet_point1_DE_FR = 54
col_bullet_point2_DE_FR = 55
col_bullet_point3_DE_FR = 56
col_bullet_point4_DE_FR = 57
col_bullet_point5_DE_FR = 58
col_parent_child_others = 37
col_parent_sku_others = 38
col_relationship_type_others = 39
col_variation_theme_others = 40
col_update_delete_others = 41
col_bullet_point1_IT_ES = 53
col_bullet_point2_IT_ES = 54
col_bullet_point3_IT_ES = 55
col_bullet_point4_IT_ES = 56
col_bullet_point5_IT_ES = 57
col_material_composition_IT_ES= 18

id_type = ws_Rawdata["B18"].value

"""Define the case"""

def Case_A (col,contents):
    ws_Temp.cell(row = cur_row + i,column = col).value = contents

def Case_B (col):

    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = item_sku
    elif i < 10:
        ws_Temp.cell(row=cur_row + i, column=col).value = item_sku + "-0" +str(i)
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = item_sku + "-" +str(i)

def Case_C (col,contents1,contents2):
    if contents2 == None:
        ws_Temp.cell(row=cur_row + i, column=col).value = contents1
    else:
        if i % 2 == 0:
            ws_Temp.cell(row=cur_row + i, column=col).value = contents1
        else:
            ws_Temp.cell(row=cur_row + i, column=col).value = contents2

def Case_D (col):
    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = "Parent"
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = "Child"

def Case_E (col,contents):

    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = None
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = contents
"""Update the file"""
counter_product = 0
for counter_product in range(0,Product_Qty):
    if ws_Rawdata.cell(row = 26, column = 2 + counter_product).value == "Uploaded":
        counter_product += 1
    else:
        item_sku = ws_Rawdata.cell(row=3, column=2 + counter_product).value
        item_name = ws_Rawdata.cell(row=4, column=2 + counter_product).value
        product_description = ws_Rawdata.cell(row=5, column=2 + counter_product).value
        generic_keywords1 = ws_Rawdata.cell(row=6, column=2 + counter_product).value
        generic_keywords2 = ws_Rawdata.cell(row=7, column=2 + counter_product).value
        bullet_point1 = ws_Rawdata.cell(row=8, column=2 + counter_product).value
        bullet_point2 = ws_Rawdata.cell(row=9, column=2 + counter_product).value
        bullet_point3 = ws_Rawdata.cell(row=10, column=2 + counter_product).value
        bullet_point4 = ws_Rawdata.cell(row=11, column=2 + counter_product).value
        bullet_point5 = ws_Rawdata.cell(row=12, column=2 + counter_product).value
        standard_price = ws_Rawdata.cell(row=13, column=2 + counter_product).value
        recommended_browse_nodes1 = ws_Rawdata.cell(row=14, column=2 + counter_product).value
        recommended_browse_nodes2 = ws_Rawdata.cell(row=15, column=2 + counter_product).value
        outer_material_type = ws_Rawdata.cell(row=16, column=2 + counter_product).value
        brand_name = ws_Rawdata.cell(row=17, column=2 + counter_product).value
        quantity = ws_Rawdata.cell(row=19, column=2 + counter_product).value
        relationship_type = ws_Rawdata.cell(row=22, column=2 + counter_product).value
        variation_theme = ws_Rawdata.cell(row=24, column=2 + counter_product).value
        update_delete = ws_Rawdata.cell(row=25, column=2 + counter_product).value
        qty_child = Child_qty[counter_product]
        if ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "NULL":
            counter_product += 1
        elif ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "UK":
            Temp = "Template_"+ws_Rawdata.cell(row = 1, column = 2 + counter_product).value
            ws_Temp = wb_updatedfile[Temp]
            cur_row = cur_row_UK
            i = 0
            for i in range (0,qty_child+1):
                Case_A(col_brand_name, brand_name)
                Case_A(col_external_product_id_type, id_type)
                Case_B(col_item_sku)
                Case_D(col_parent_child_UK)
                Case_E(col_parent_sku_UK, item_sku)
                Case_E(col_quantity, quantity)
                Case_E(col_relationship_type_UK, relationship_type)
                Case_A(col_item_name, item_name)
                Case_A(col_product_description_UK, product_description)
                Case_A(col_bullet_point1_UK, bullet_point1)
                Case_A(col_bullet_point2_UK, bullet_point2)
                Case_A(col_bullet_point3_UK, bullet_point3)
                Case_A(col_bullet_point4_UK, bullet_point4)
                Case_A(col_bullet_point5_UK, bullet_point5)
                Case_C(col_generic_keywords1_UK, generic_keywords1, generic_keywords2)
                Case_C(col_recommended_browse_nodes, recommended_browse_nodes1, recommended_browse_nodes2)
                Case_A(col_outer_material_type, outer_material_type)
                Case_A(col_material_composition_UK_DE_FR, outer_material_type)
                Case_A(col_variation_theme_UK, variation_theme)
                Case_A(col_update_delete_UK, update_delete)
                Case_E(col_standard_price, standard_price)
            cur_row_UK = cur_row + qty_child +1
            ws_Rawdata.cell(row=26, column=2 + counter_product).value = "Uploaded"
            counter_product += 1
        elif ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "DE" or ws_Rawdata.cell(row=1, column=2 + counter_product).value == "FR":
            Temp = "Template_"+ws_Rawdata.cell(row = 1, column = 2 + counter_product).value
            ws_Temp = wb_updatedfile[Temp]
            if ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "DE":
                cur_row = cur_row_DE
            else:
                cur_row = cur_row_FR
            i = 0
            for i in range (0,qty_child+1):
                Case_A(col_brand_name, brand_name)
                Case_A(col_external_product_id_type, id_type)
                Case_B(col_item_sku)
                Case_D(col_parent_child_others)
                Case_E(col_parent_sku_others, item_sku)
                Case_E(col_quantity, quantity)
                Case_E(col_relationship_type_others, relationship_type)
                Case_A(col_item_name, item_name)
                Case_A(col_product_description_others, product_description)
                Case_A(col_bullet_point1_DE_FR, bullet_point1)
                Case_A(col_bullet_point2_DE_FR, bullet_point2)
                Case_A(col_bullet_point3_DE_FR, bullet_point3)
                Case_A(col_bullet_point4_DE_FR, bullet_point4)
                Case_A(col_bullet_point5_DE_FR, bullet_point5)
                Case_C(col_generic_keywords1_others, generic_keywords1, generic_keywords2)
                Case_C(col_recommended_browse_nodes, recommended_browse_nodes1, recommended_browse_nodes2)
                Case_A(col_outer_material_type, outer_material_type)
                Case_A(col_material_composition_UK_DE_FR, outer_material_type)
                Case_A(col_variation_theme_others, variation_theme)
                Case_A(col_update_delete_others, update_delete)
                Case_E(col_standard_price, standard_price)
            if ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "DE":
                cur_row_DE = cur_row + qty_child + 1
            else:
                cur_row_FR = cur_row + qty_child + 1
            ws_Rawdata.cell(row=26, column=2 + counter_product).value = "Uploaded"
            counter_product += 1
        else:
            Temp = "Template_"+ws_Rawdata.cell(row = 1, column = 2 + counter_product).value
            ws_Temp = wb_updatedfile[Temp]
            if ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "IT":
                cur_row = cur_row_IT
            else:
                cur_row = cur_row_ES
            i = 0
            for i in range (0,qty_child+1):
                Case_A(col_brand_name, brand_name)
                Case_A(col_external_product_id_type, id_type)
                Case_B(col_item_sku)
                Case_D(col_parent_child_others)
                Case_E(col_parent_sku_others, item_sku)
                Case_E(col_quantity, quantity)
                Case_E(col_relationship_type_others, relationship_type)
                Case_A(col_item_name, item_name)
                Case_A(col_product_description_others, product_description)
                Case_A(col_bullet_point1_IT_ES, bullet_point1)
                Case_A(col_bullet_point2_IT_ES, bullet_point2)
                Case_A(col_bullet_point3_IT_ES, bullet_point3)
                Case_A(col_bullet_point4_IT_ES, bullet_point4)
                Case_A(col_bullet_point5_IT_ES, bullet_point5)
                Case_C(col_generic_keywords1_others, generic_keywords1, generic_keywords2)
                Case_C(col_recommended_browse_nodes, recommended_browse_nodes1, recommended_browse_nodes2)
                Case_A(col_outer_material_type, outer_material_type)
                Case_A(col_material_composition_IT_ES, outer_material_type)
                Case_A(col_variation_theme_others, variation_theme)
                Case_A(col_update_delete_others, update_delete)
                Case_E(col_standard_price, standard_price)
            if ws_Rawdata.cell(row = 1, column = 2 + counter_product).value == "IT":
                cur_row_IT = cur_row + qty_child + 1
            else:
                cur_row_ES = cur_row + qty_child + 1
            ws_Rawdata.cell(row=26, column=2 + counter_product).value = "Uploaded"
            counter_product += 1

"""Save the file"""

wb_updatedfile.save('D:\Python\Python\Ivy part time\\Updated file.xlsx')
os.startfile(r'D:\Python\Python\Ivy part time\\Updated file.xlsx')
wb_rawdata.save('D:\Python\Python\Ivy part time\\Raw data.xlsx')
os.startfile(r'D:\Python\Python\Ivy part time\\Raw data.xlsx')

