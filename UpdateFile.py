from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb_rawdata = load_workbook('E:\Python\Ivy part time\\Raw data.xlsx')
wb_updatedfile = load_workbook('E:\Python\Ivy part time\\Updated file.xlsx')
ws_Temp = wb_updatedfile["Template"]
ws_Rawdata = wb_rawdata["Rawdata"]

"""list out the product & child_qty"""

Child_qty = []
for cell in ws_Rawdata["1"]:
    col = cell.column
    Child_qty.append(ws_Rawdata.cell(row = 1,column = col).value)

del Child_qty[0]
del Child_qty[0]


"""define the column of these info """

cur_row = 4
col_item_sku = 2
col_brand_name = 3
col_external_product_id_type = 5
col_item_name = 6
col_quantity = 27
col_parent_child = 32
col_parent_sku = 33
col_relationship_type = 34
col_product_description = 37
col_generic_keywords1 = 43
col_bullet_point1 = 49
col_bullet_point2 = 50
col_bullet_point3 = 51
col_bullet_point4 = 52
col_bullet_point5 = 53
col_recommended_browse_nodes = 7
col_outer_material_type = 8
col_material_composition = 13
col_variation_theme = 35
col_update_delete = 36

brand_name = ws_Rawdata["C3"].value
id_type = ws_Rawdata["C4"].value
counter_product = 0

"""Define the case"""

def Case_A (col,contents):
    ws_Temp.cell(row = cur_row + i,column = col).value = contents

def Case_B (col):

    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = parent_sku
    elif i < 10:
        ws_Temp.cell(row=cur_row + i, column=col).value = parent_sku + "-0" +str(i)
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = parent_sku + "-" +str(i)

def Case_C (col,contents1,contents2):
    if contents2 == None:
        ws_Temp.cell(row=cur_row + i, column=col).value = contents1
    else:
        if i % 2 == 0:
            ws_Temp.cell(row=cur_row + i, column=col).value = contents1
        else:
            ws_Temp.cell(row=cur_row + i, column=col).value = contents2


def Case_D (col):
    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = "Parent"
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = "Child"

def Case_E (col,contents):

    if i == 0:
        ws_Temp.cell(row=cur_row + i, column=col).value = None
    else:
        ws_Temp.cell(row=cur_row + i, column=col).value = contents
"""Update the file"""

for product in Child_qty:
    counter_product += 1
    parent_sku = ws_Rawdata.cell(row = 2, column = 2 + counter_product).value
    quantity = ws_Rawdata.cell(row = 6, column = 2 + counter_product).value
    relationship_type = "Variation"
    item_name = ws_Rawdata.cell(row = 5, column = 2 + counter_product).value
    product_description = ws_Rawdata.cell(row = 10, column = 2 + counter_product).value
    bullet_point1 = ws_Rawdata.cell(row = 13, column = 2 + counter_product).value
    bullet_point2 = ws_Rawdata.cell(row = 14, column = 2 + counter_product).value
    bullet_point3 = ws_Rawdata.cell(row = 15, column = 2 + counter_product).value
    bullet_point4 = ws_Rawdata.cell(row = 16, column = 2 + counter_product).value
    bullet_point5 = ws_Rawdata.cell(row = 17, column = 2 + counter_product).value
    generic_keywords1 = ws_Rawdata.cell(row = 11, column = 2 + counter_product).value
    generic_keywords2 = ws_Rawdata.cell(row = 12, column = 2 + counter_product).value
    recommended_browse_nodes1 = ws_Rawdata.cell(row = 18, column = 2 + counter_product).value
    recommended_browse_nodes2 = ws_Rawdata.cell(row = 19, column = 2 + counter_product).value
    outer_material_type = ws_Rawdata.cell(row = 20, column = 2 + counter_product).value
    variation_theme = ws_Rawdata.cell(row = 22, column = 2 + counter_product).value
    update_delete = ws_Rawdata.cell(row = 23, column = 2 + counter_product).value
    for i in range(0, product+1):
        Case_A (col_brand_name,brand_name)
        Case_A (col_external_product_id_type,id_type)
        Case_B (col_item_sku)
        Case_D (col_parent_child)
        Case_E (col_parent_sku,parent_sku)
        Case_E (col_quantity,quantity)
        Case_E (col_relationship_type,relationship_type)
        Case_A (col_item_name, item_name)
        Case_A (col_product_description, product_description)
        Case_A (col_bullet_point1, bullet_point1)
        Case_A (col_bullet_point2, bullet_point2)
        Case_A (col_bullet_point3, bullet_point3)
        Case_A (col_bullet_point4, bullet_point4)
        Case_A (col_bullet_point5, bullet_point5)
        Case_C (col_generic_keywords1, generic_keywords1,generic_keywords2)
        Case_C (col_recommended_browse_nodes, recommended_browse_nodes1, recommended_browse_nodes2)
        Case_A (col_outer_material_type, outer_material_type)
        Case_A (col_material_composition, outer_material_type)
        Case_A (col_variation_theme, variation_theme)
        Case_A (col_update_delete, update_delete)

    cur_row += product + 1

"""Save the file"""

wb_updatedfile.save('E:\Python\Ivy part time\\Updated file.xlsx')
os.startfile(r'E:\Python\Ivy part time\\Updated file.xlsx')


